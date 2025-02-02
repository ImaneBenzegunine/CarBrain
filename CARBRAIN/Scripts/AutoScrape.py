import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import os
from path import PATH_AUTO_XLSX, PATH_AUTO_CSV
import csv

class CarScraper:
    def __init__(self, headless=True):
        self.options = uc.ChromeOptions()
        if headless:
            self.options.add_argument('--headless')  
        self.options.add_argument('--disable-gpu')
        self.options.add_argument('--no-sandbox')
        self.options.add_argument('--disable-dev-shm-usage')
        self.options.add_argument('--remote-debugging-port=9222')  # Add this for debugging
        self.driver = uc.Chrome(options=self.options)
        
    def open_page(self, url):
        """Open the page and wait for the content to load."""
        self.driver.get(url)
        WebDriverWait(self.driver, 30).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, "ListItem_wrapper__TxHWu"))
        )
    
    def extract_car_links(self, page_number):
        """Extract car links from a given page."""
        website = f"https://www.autoscout24.be/fr/lst?atype=C&cy=B&desc=0&page={page_number}&search_id=a5i2bcp2m1&sort=standard&source=listpage_pagination&ustate=N%2CU"
        self.open_page(website)
        cars_sections = self.driver.find_elements(By.CLASS_NAME, "ListItem_wrapper__TxHWu")
        links = []
        for car in cars_sections:
            try:
                link_car = WebDriverWait(car, 30).until(
                    EC.presence_of_element_located((By.XPATH, './/a'))
                )
                car_url = link_car.get_attribute('href')
                links.append(car_url)
            except Exception as e:
                print(f"Error while retrieving car link: {e}")
        return links
    
    def extract_car_details(self, car_url):
        """Extract car details from a given car page."""
        self.driver.get(car_url)

        # Wait for the basic details section to load
        WebDriverWait(self.driver, 30).until(           
            EC.presence_of_element_located((By.XPATH, '//*[@id="basic-details-section"]/div/div[2]/dl'))
        )
        donnebase = self.driver.find_element(By.XPATH, '//*[@id="basic-details-section"]/div/div[2]/dl')
        Keys = donnebase.find_elements(By.TAG_NAME, 'dt')
        Vals = donnebase.find_elements(By.TAG_NAME, 'dd')

        # Extract car name
        WebDriverWait(self.driver, 30).until(            
            EC.presence_of_element_located((By.CLASS_NAME, "StageTitle_boldClassifiedInfo__sQb0l"))
        )   
        CarName = self.driver.find_element(By.CLASS_NAME, "StageTitle_boldClassifiedInfo__sQb0l")

        # Extract car model
        WebDriverWait(self.driver, 30).until(         
            EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/div/div/main/div[3]/div[2]/div[1]/div[2]/h1/div[1]/span[2]'))
        )
        CarModele = self.driver.find_element(By.XPATH, '//*[@id="__next"]/div/div/main/div[3]/div[2]/div[1]/div[2]/h1/div[1]/span[2]')

        # Extract car price
        WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'PriceInfo_price__XU0aF'))
        )
        CarPrice = self.driver.find_element(By.CLASS_NAME, 'PriceInfo_price__XU0aF')

        # Extract car condition
        WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/div/div/main/div[3]/div[2]/div[3]/div[2]/div/div'))
        )
        CarEtat = self.driver.find_element(By.XPATH, '//*[@id="__next"]/div/div/main/div[3]/div[2]/div[3]/div[2]/div/div')

        #===> CarsInfoSection1
        WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'StageArea_overviewContainer__UyZ9n'))
        )
        CarsInfo = self.driver.find_element(By.CLASS_NAME, 'StageArea_overviewContainer__UyZ9n')
        CarInfoDet = CarsInfo.find_elements(By.CLASS_NAME, "VehicleOverview_itemText__AI4dA")
        CarInfoDetMore = [detail.text for detail in CarInfoDet]


        WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="technical-details-section"]/div/div[2]/dl'))
        )
        MoreCaracs = self.driver.find_element(By.XPATH, '//*[@id="technical-details-section"]/div/div[2]/dl')
        MoreCaracs = MoreCaracs.find_elements(By.TAG_NAME, 'dd')
        chaine = " ".join(s.text for s in MoreCaracs) + " "


        donn = {}
        for key, val in zip(Keys, Vals):
            if key.text == "Sièges" or key.text == "Portes":
                donn[key.text] = val.text
        return [CarName.text, CarModele.text, CarPrice.text, CarEtat.text, CarInfoDetMore[0], CarInfoDetMore[1], CarInfoDetMore[2], CarInfoDetMore[3], CarInfoDetMore[4],
                 CarInfoDetMore[5], donn.get('Sièges', 'N/A'), donn.get('Portes', 'N/A'), chaine]
    
    def verify_xlsx_file(self, path = PATH_AUTO_XLSX) :
        if os.path.exists(path):
            try:
                wb = load_workbook(path)  # Essayer de charger le fichier
                ws = wb.active
            except Exception as e:
                print(f"Erreur lors de l'ouverture du fichier : {e}")
                print("Création d'un nouveau fichier Excel...")
                os.remove(path)  # Supprimer le fichier corrompu
                wb = Workbook()
                ws = wb.active
                ws.append(["car_name", "carModele", "CarPrice", "Etat", "Milieage", "Tran", 
                           "Annee", "CarCar", "CarPui", "CarVen", "sieges", "seats", "Options", "link"])
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["car_name", "carModele", "CarPrice", "Etat", "Milieage", "Tran", 
                           "Annee", "CarCar", "CarPui", "CarVen", "technical_details", "seats", "Options", "link"])
            wb.save(path)

    def verify_csv_file(self, path = PATH_AUTO_CSV):
        """Verify if the CSV file exists and create it with headers if necessary."""
        if not os.path.exists(path):
            with open(path, mode='w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(["car_name", "carModele", "CarPrice", "Etat", "Milieage", "Tran", 
                                "Annee", "CarCar", "CarPui", "CarVen", "technical_details", "seats", "Options", "link"])


    def save_car_details(self, car_url, path_xlsx = PATH_AUTO_XLSX, path_csv = PATH_AUTO_CSV) :
        self.verify_xlsx_file()
        self.verify_csv_file()
        carlist = self.extract_car_details(car_url)
        wb = load_workbook(PATH_AUTO_XLSX)
        ws = wb.active
        ws.append(carlist + [car_url])
        wb.save(PATH_AUTO_XLSX)
        with open(PATH_AUTO_CSV, mode='a', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(carlist + [car_url])
        
    def close(self):
        """Close the browser."""
        self.driver.quit()

def main():
    scraper = CarScraper()

    try:
        i = 1
        while True:
            car_links = scraper.extract_car_links(i)
            if not car_links:  
                print(f"No cars found on page {i}. Stopping.")
                break
            
            print(f"Number of cars on page {i}: {len(car_links)}")
            
            for car_url in car_links:
                try:
                    scraper.save_car_details(car_url)
                except Exception as e:
                    print(f"Error while retrieving car information: {e}")
            i += 1

    except Exception as e:
        print(f"General error: {e}")
    finally:
        scraper.close()

if __name__ == "__main__":
    main()